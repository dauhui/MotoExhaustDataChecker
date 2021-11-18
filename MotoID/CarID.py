# Motorbike scheduled maintenance checker & writer
# by Rex Yang 10/15/2021

from selenium import webdriver
# import xlrd # use this if .xls
import openpyxl # use this if .xlsx
print("注意，請使用備份檔，避免原檔案損壞")
print("請將欲使用之Excel檔改名為test")
print("並確認副檔名為.xlsx(新版Excel)")
input("確認後，請按任意鍵繼續")
workbook = openpyxl.load_workbook('test.xlsx')
worksheet = workbook.worksheets[0]
IDnum = worksheet.max_row - 1 # assume the first i rows aren't ID

print("請輸入車牌所在欄位，A為1、B為2...、T為20，依此類推")
IDrow = input("欄位: ")
CarID = list(range(IDnum))  # create list with selected length
for i in range(2, IDnum + 2):
    # CHANGE HERE
    CarID[i-2] = worksheet.cell(i,IDrow).value  # store ID in to list from excel

# set specific item id from website
input_id = 'ctl00_MainContent_txtCarNo'
submit_id = 'ctl00_MainContent_btnQuery'
isdone_id = 'ctl00_MainContent_lbIsDone'

def inputCarID(driver, InputID, element_class):
    text_answers = [str(InputID)]
    text_questions = driver.find_elements_by_id(element_class)
    for a,q in zip(text_answers,text_questions):
        q.send_keys(a)
    return driver

def submit(driver, element_class):
    driver.find_element_by_id(element_class).click()
    return driver

def isdone(element_class):
    tf = driver.find_element_by_id(element_class).text
    return (driver, tf)

def clear():
    driver.find_element_by_id(input_id).clear()
    return driver

url = "https://www.motorim.org.tw/query/query_check.aspx"
driver = webdriver.Edge(executable_path="./msedgedriver")
driver.get(url)

Outrow = input("請輸入欲顯示結果之欄位(英文大寫)")
for i in range(0, IDnum):
    driver = inputCarID(driver, CarID[i], input_id)
    driver = submit(driver, submit_id)
    (driver, tf) = isdone(isdone_id)
    # CHANGE HERE
    worksheet[Outrow + str(i + 2)] = tf
    driver = clear()
workbook.save('test.xlsx')